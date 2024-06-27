import openpyxl


def get_Row_count(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet.max_row)

def get_Coloumn_count(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet.max_column)

def read_data(file,sheetname,rownum,colomn_no):
    workbook =openpyxl.load_workbook(file)
